from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import get_data_dir
from database import get_all_companies, get_contacts_by_company, init_database


HEADERS = [
    "公司 ID",
    "公司名称",
    "官网",
    "国家/地区",
    "行业",
    "状态",
    "置信度",
    "判断原因",
    "联系人邮箱",
    "邮箱类型",
    "来源网址",
    "发送时间",
    "最后错误",
    "创建时间",
]


def export_to_excel(output_path="data/leads.xlsx") -> str:
    init_database()
    out_path = Path(output_path)
    if not out_path.is_absolute():
        out_path = get_data_dir().parent / out_path
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "潜在客户"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for company in get_all_companies():
        contacts = get_contacts_by_company(company["id"])
        if not contacts:
            contacts = [None]
        for contact in contacts:
            ws.append(
                [
                    company["id"],
                    company["company_name"],
                    company["website"],
                    company["country"],
                    company["industry"],
                    company["status"],
                    company["confidence"],
                    company["reason"],
                    contact["email"] if contact else "",
                    contact["email_type"] if contact else "",
                    contact["source_url"] if contact else "",
                    company["sent_at"],
                    company["last_error"],
                    company["created_at"],
                ]
            )

    for column_cells in ws.columns:
        max_len = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column].width = min(max(max_len + 2, 12), 60)

    ws.freeze_panes = "A2"
    wb.save(out_path)
    return str(out_path)
